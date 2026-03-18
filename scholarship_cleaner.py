#!/usr/bin/env python3
"""
ScholarBot Pro — Scholarship Cleaner & Deduplicator
Reads all 3 Excel source files + the 30 built-in scholarships,
normalizes to a canonical schema, deduplicates, and outputs a clean master CSV.
"""

import csv
import hashlib
import re
import os
from datetime import datetime, date
from difflib import SequenceMatcher
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================
PROJECT_FILES = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "mnt", ".projects", "019c4f5c-2c27-774c-a9e0-158925d25163", "files"
)
EXCEL_FILES = [
    "9dd796c4-40ff-48da-b5bd-4585393b22f7.xlsx",
    "b7eb9051-edb0-4ee1-aee1-10d7abb5f17f.xlsx",
    "dd918db9-7ae8-4f4e-a149-1ad5165294c2.xlsx",
]
OUTPUT_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scholarship_master_clean.csv")
SIMILARITY_THRESHOLD = 0.85  # Names above this are considered duplicates


# ============================================================
# 30 BUILT-IN VERIFIED SCHOLARSHIPS
# ============================================================
BUILTIN_SCHOLARSHIPS = [
    {"name":"Gates Scholarship","criteria":"High school seniors from minority backgrounds (African American, Hispanic, Asian/Pacific Islander, Native American). Pell-eligible. Must demonstrate leadership and academic excellence. 3.3+ GPA on 4.0 scale. U.S. citizen, national, or permanent resident.","link":"https://www.thegatesscholarship.org/","deadline":"2026-09-15","amount":"Full Tuition","needBased":"Y"},
    {"name":"Ron Brown Scholar Program","criteria":"African American high school seniors. Must demonstrate academic excellence, leadership, and community service. U.S. citizen or permanent resident. Financial need considered.","link":"https://ronbrown.org/ron-brown-scholarship/","deadline":"2026-12-01","amount":"$40,000","needBased":"Y"},
    {"name":"Coca-Cola Scholars Foundation","criteria":"High school seniors with leadership in school and community. U.S. citizens, nationals, permanent residents, refugees, or asylees. Must be eligible for federal financial aid. Achievement-based.","link":"https://www.coca-colascholarsfoundation.org/apply/","deadline":"2026-09-30","amount":"$20,000","needBased":""},
    {"name":"Dell Scholars Program","criteria":"Must participate in an approved college readiness program. Demonstrate need for financial assistance. GPA of 2.4+. U.S. citizen or permanent resident. Must be a current high school senior.","link":"https://www.dellscholars.org/","deadline":"2026-12-01","amount":"$20,000","needBased":"Y"},
    {"name":"QuestBridge National College Match","criteria":"High-achieving low-income students. Typically household income under $65,000. Strong academics. High school seniors applying to partner colleges.","link":"https://www.questbridge.org/","deadline":"2026-09-26","amount":"Full Ride","needBased":"Y"},
    {"name":"Elks Most Valuable Student Scholarship","criteria":"U.S. citizen high school senior. Judged on scholarship, leadership, financial need. Must plan to pursue a four-year degree.","link":"https://www.elks.org/scholars/scholarships/mvs.cfm","deadline":"2026-11-05","amount":"$12,500","needBased":"Y"},
    {"name":"Burger King Scholars Program","criteria":"High school seniors in U.S., Canada, Puerto Rico, or Guam. GPA 2.0+. Demonstrate financial need, work experience, community involvement.","link":"https://burgerking.scholarsapply.org/","deadline":"2026-12-15","amount":"$1,000-$60,000","needBased":"Y"},
    {"name":"Cameron Impact Scholarship","criteria":"High school seniors. Demonstrated academic achievement, community involvement, and leadership. U.S. citizens. Plan to attend four-year institution.","link":"https://www.bryancameroneducationfoundation.org/","deadline":"2026-09-14","amount":"Full Tuition","needBased":""},
    {"name":"Daniels Fund Scholarship","criteria":"Graduating high school seniors from CO, NM, UT, WY. Demonstrate strength of character, leadership, community service. Financial need.","link":"https://www.danielsfund.org/scholarships","deadline":"2026-11-15","amount":"Full Tuition","needBased":"Y"},
    {"name":"UNCF Scholarships","criteria":"Underrepresented minority students. Multiple scholarship programs available year-round. Must attend an HBCU or other accredited institution.","link":"https://uncf.org/scholarships","deadline":"Varies","amount":"Varies","needBased":"Y"},
    {"name":"Hispanic Scholarship Fund","criteria":"Of Hispanic heritage. U.S. citizen, permanent resident, or DACA eligible. Minimum 3.0 GPA. Plan to enroll full-time in accredited institution.","link":"https://www.hsf.net/scholarship","deadline":"2026-02-15","amount":"$500-$5,000","needBased":""},
    {"name":"Asian & Pacific Islander American Scholarship (APIASF)","criteria":"Asian American or Pacific Islander ethnicity. 2.7+ GPA. U.S. citizen, national, permanent resident, or citizen of Freely Associated States. Financial need.","link":"https://apiascholars.org/","deadline":"2026-01-11","amount":"Up to $20,000","needBased":"Y"},
    {"name":"Equitable Excellence Scholarship","criteria":"High school senior. U.S. citizen or legal resident. 2.5+ GPA. Demonstrate leadership, determination, and resilience.","link":"https://equitable.com/foundation/equitable-excellence-scholarship","deadline":"2026-12-18","amount":"$5,000/yr renewable","needBased":""},
    {"name":"Horatio Alger Scholarship","criteria":"High school senior. Demonstrated financial need (family income under $55,000). Minimum 2.0 GPA. Involvement in co-curricular and community activities. U.S. citizen.","link":"https://scholars.horatioalger.org/","deadline":"2026-10-25","amount":"$25,000","needBased":"Y"},
    {"name":"Jack Kent Cooke Foundation College Scholarship","criteria":"High school senior with financial need (family income under $95,000). 3.5+ unweighted GPA. Standardized test scores. U.S. citizen or permanent resident.","link":"https://www.jkcf.org/our-scholarships/","deadline":"2026-11-18","amount":"Up to $55,000/yr","needBased":"Y"},
    {"name":"Posse Foundation Scholarship","criteria":"Must be nominated by high school. Urban public high school students with extraordinary leadership potential. Full tuition at partner colleges.","link":"https://www.possefoundation.org/","deadline":"Nomination Only","amount":"Full Tuition","needBased":""},
    {"name":"Regeneron Science Talent Search","criteria":"High school seniors in the U.S. Must submit original research project in science, math, or engineering.","link":"https://www.societyforscience.org/regeneron-sts/","deadline":"2026-11-12","amount":"Up to $250,000","needBased":""},
    {"name":"National Merit Scholarship","criteria":"U.S. high school students. Based on PSAT/NMSQT scores taken in junior year. Must be enrolled or plan to enroll full-time in college.","link":"https://www.nationalmerit.org/","deadline":"2026-10-01","amount":"$2,500+","needBased":""},
    {"name":"Cobell Scholarship (Native American)","criteria":"Must be enrolled member of a federally recognized tribe. Undergraduate or graduate student. Financial need demonstrated.","link":"https://cobellscholar.org/","deadline":"2026-01-31","amount":"Up to $5,000","needBased":"Y"},
    {"name":"NAACP Scholarships","criteria":"African American students. Must be current NAACP member. Academic merit and financial need considered.","link":"https://naacp.org/find-resources/scholarships","deadline":"Varies","amount":"Varies","needBased":"Y"},
    {"name":"Dream.US Scholarship (DREAMers)","criteria":"DACA or TPS recipients. First-time college students or community college transfers. Financial need. 2.5+ GPA. Must attend a partner college.","link":"https://www.thedream.us/","deadline":"2026-02-28","amount":"Up to $33,000","needBased":"Y"},
    {"name":"GE-Reagan Foundation Scholarship","criteria":"High school senior. U.S. citizen. Demonstrate leadership, drive, integrity, and citizenship. 3.0+ GPA.","link":"https://www.reaganfoundation.org/education/scholarship-programs/","deadline":"2026-01-05","amount":"$10,000/yr renewable","needBased":""},
    {"name":"Amazon Future Engineer Scholarship","criteria":"High school senior planning to study computer science. Financial need. Participation in STEM activities. Includes paid internship.","link":"https://www.amazonfutureengineer.com/scholarships","deadline":"2026-01-20","amount":"$40,000","needBased":"Y"},
    {"name":"Buick Achievers Scholarship","criteria":"High school senior or current undergraduate. Plan to major in a STEM field. Financial need. Leadership and community involvement.","link":"https://www.buickachievers.com/","deadline":"2026-02-28","amount":"$25,000","needBased":"Y"},
    {"name":"Davidson Fellows Scholarship","criteria":"Students 18 or under. Must complete a significant project in STEM, literature, music, philosophy. U.S. citizen or permanent resident.","link":"https://www.davidsongifted.org/gifted-programs/fellows-scholarship/","deadline":"2026-02-11","amount":"$10,000-$50,000","needBased":""},
    {"name":"Prudential Emerging Visionaries","criteria":"Ages 14-18. Must have created a financial or societal solution for your community. U.S. residents.","link":"https://www.prudential.com/emerging-visionaries","deadline":"2026-11-01","amount":"Up to $15,000","needBased":""},
    {"name":"Taco Bell Live Mas Scholarship","criteria":"Ages 16-26. Must be pursuing education at an accredited institution in the U.S. Based on passion and innovation, not just grades.","link":"https://www.tacobellfoundation.org/live-mas-scholarship/","deadline":"2026-01-24","amount":"$5,000-$25,000","needBased":""},
    {"name":"Jackie Robinson Foundation Scholarship","criteria":"Minority high school senior with leadership potential. SAT/ACT scores considered. Financial need demonstrated. Must be U.S. citizen.","link":"https://www.jackierobinson.org/apply/","deadline":"2026-02-01","amount":"Up to $30,000","needBased":"Y"},
    {"name":"Foot Locker Foundation-UNCF Scholarship","criteria":"Students attending a UNCF member HBCU. Minimum 2.5 GPA. Demonstrate financial need.","link":"https://uncf.org/scholarships","deadline":"2026-04-10","amount":"$5,000","needBased":"Y"},
    {"name":"TMCF Coca-Cola First Generation HBCU Scholarship","criteria":"First-generation college student. Graduating high school senior. Enrolling full-time at a TMCF member HBCU. Financial need.","link":"https://tmcf.org/","deadline":"2026-05-01","amount":"$5,000","needBased":"Y"},
]


def generate_id(name, link=""):
    """Generate a stable short hash ID from name + link."""
    raw = (name.strip().lower() + link.strip().lower()).encode("utf-8")
    return hashlib.md5(raw).hexdigest()[:8]


def normalize_name(name):
    """Normalize scholarship name for comparison."""
    if not name:
        return ""
    name = name.strip()
    # Remove common suffixes/noise
    name = re.sub(r'\s*\(.*?\)\s*$', '', name)  # Remove trailing parenthetical
    name = re.sub(r'\s+', ' ', name)  # Normalize whitespace
    return name


def normalize_deadline(raw):
    """Try to parse deadline into YYYY-MM-DD or return descriptive string."""
    if not raw:
        return "Varies"
    raw = str(raw).strip()

    # Already ISO format
    if re.match(r'^\d{4}-\d{2}-\d{2}$', raw):
        return raw

    # Common descriptive deadlines
    lower = raw.lower()
    if any(w in lower for w in ["varies", "ongoing", "rolling", "multiple", "year-round", "nomination", "closed"]):
        return raw.title()

    # Try parsing datetime objects (from Excel)
    if hasattr(raw, 'strftime'):
        return raw.strftime('%Y-%m-%d')

    # Try common date formats
    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%B %d, %Y', '%b %d, %Y', '%d-%b-%Y']:
        try:
            return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            continue

    return raw


def normalize_amount(raw):
    """Clean up amount field."""
    if not raw:
        return "Varies"
    raw = str(raw).strip()
    if not raw or raw.lower() in ["varies", "unknown", "tbd", "n/a", "na"]:
        return "Varies"
    return raw


def normalize_need_based(raw):
    """Normalize need-based field to 'Y' or ''."""
    if not raw:
        return ""
    raw = str(raw).strip().upper()
    if raw in ["Y", "YES", "TRUE", "1", "NEED", "NEED-BASED", "NEED BASED"]:
        return "Y"
    return ""


def is_similar(name1, name2, threshold=SIMILARITY_THRESHOLD):
    """Check if two scholarship names are similar enough to be duplicates."""
    n1 = normalize_name(name1).lower()
    n2 = normalize_name(name2).lower()
    if not n1 or not n2:
        return False
    # Exact match after normalization
    if n1 == n2:
        return True
    # One contains the other
    if n1 in n2 or n2 in n1:
        return True
    # Fuzzy match
    return SequenceMatcher(None, n1, n2).ratio() >= threshold


def extract_from_excel(filepath):
    """Extract scholarships from an Excel file, handling multiple sheet layouts."""
    scholarships = []
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR reading {filepath}: {e}")
        return scholarships

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Find header row (first row with recognizable column names)
        header = None
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            cells = [str(c).strip().lower() if c else "" for c in row]
            if any(k in " ".join(cells) for k in ["name", "scholarship", "criteria", "link", "deadline", "amount"]):
                header = cells
                header_idx = i
                break

        if not header:
            continue

        # Map columns
        col_map = {}
        for j, h in enumerate(header):
            if not h:
                continue
            if "name" in h or "scholarship" in h:
                col_map.setdefault("name", j)
            elif "criteria" in h or "description" in h or "detail" in h or "required" in h:
                col_map.setdefault("criteria", j)
            elif "link" in h or "url" in h or "info" in h or "application" in h:
                col_map.setdefault("link", j)
            elif "deadline" in h or "due" in h or "date" in h:
                col_map.setdefault("deadline", j)
            elif "amount" in h or "award" in h or "max" in h:
                col_map.setdefault("amount", j)
            elif "need" in h:
                col_map.setdefault("needBased", j)
            elif "functional" in h:
                col_map.setdefault("functional", j)

        if "name" not in col_map:
            continue

        # Extract rows
        for row in rows[header_idx + 1:]:
            if not row or not row[col_map["name"]]:
                continue
            name = str(row[col_map["name"]]).strip()
            if not name or name.lower() in ["name", "scholarship name", "none", ""]:
                continue

            entry = {
                "name": name,
                "criteria": str(row[col_map.get("criteria", -1)] or "").strip() if "criteria" in col_map else "",
                "link": str(row[col_map.get("link", -1)] or "").strip() if "link" in col_map else "",
                "deadline": normalize_deadline(row[col_map.get("deadline", -1)]) if "deadline" in col_map else "Varies",
                "amount": normalize_amount(row[col_map.get("amount", -1)]) if "amount" in col_map else "Varies",
                "needBased": normalize_need_based(row[col_map.get("needBased", -1)]) if "needBased" in col_map else "",
                "source_sheet": f"{os.path.basename(filepath)}:{sheet_name}",
            }

            # Skip entries marked as non-functional
            if "functional" in col_map:
                func_val = str(row[col_map["functional"]] or "").strip().lower()
                if func_val in ["no", "false", "0", "expired", "closed"]:
                    continue

            # Only keep if name is substantive
            if len(name) > 3:
                entry["id"] = generate_id(name, entry["link"])
                scholarships.append(entry)

        print(f"  Sheet '{sheet_name}': extracted {len([r for r in rows[header_idx+1:] if r and r[col_map.get('name',0)]])} rows")

    wb.close()
    return scholarships


def deduplicate(scholarships):
    """Remove duplicates based on name similarity and URL matching."""
    unique = []
    seen_names = []
    seen_links = set()

    for s in scholarships:
        name = s["name"]
        link = (s.get("link") or "").strip().lower().rstrip("/")

        # Check URL duplicate
        if link and link in seen_links:
            continue

        # Check name similarity
        is_dup = False
        for existing_name in seen_names:
            if is_similar(name, existing_name):
                is_dup = True
                break

        if not is_dup:
            unique.append(s)
            seen_names.append(name)
            if link:
                seen_links.add(link)

    return unique


def main():
    print("=" * 60)
    print("ScholarBot Pro — Scholarship Cleaner & Deduplicator")
    print("=" * 60)

    all_scholarships = []

    # 1. Load built-in verified scholarships
    print(f"\n[1] Loading {len(BUILTIN_SCHOLARSHIPS)} built-in verified scholarships...")
    for s in BUILTIN_SCHOLARSHIPS:
        s["id"] = generate_id(s["name"], s.get("link", ""))
        s["source_sheet"] = "built-in-verified"
        all_scholarships.append(s)

    # 2. Load from Excel files
    print(f"\n[2] Reading {len(EXCEL_FILES)} Excel source files...")
    files_dir = PROJECT_FILES
    if not os.path.exists(files_dir):
        # Try relative path from script location
        files_dir = os.path.join(os.path.dirname(__file__), "..", "mnt", ".projects",
                                  "019c4f5c-2c27-774c-a9e0-158925d25163", "files")
    if not os.path.exists(files_dir):
        print(f"  WARNING: Project files directory not found at {files_dir}")
        print(f"  Trying alternate paths...")
        # Try absolute path
        files_dir = "/sessions/vibrant-inspiring-gates/mnt/.projects/019c4f5c-2c27-774c-a9e0-158925d25163/files"

    for fname in EXCEL_FILES:
        fpath = os.path.join(files_dir, fname)
        if os.path.exists(fpath):
            print(f"\n  Reading: {fname}")
            entries = extract_from_excel(fpath)
            print(f"  Total extracted: {len(entries)}")
            all_scholarships.extend(entries)
        else:
            print(f"  SKIP: {fname} not found at {fpath}")

    print(f"\n[3] Total raw entries: {len(all_scholarships)}")

    # 3. Sort by quality (built-in first, then by criteria length)
    all_scholarships.sort(key=lambda s: (
        0 if s.get("source_sheet") == "built-in-verified" else 1,
        -len(s.get("criteria", "")),
    ))

    # 4. Deduplicate
    print("\n[4] Deduplicating...")
    unique = deduplicate(all_scholarships)
    removed = len(all_scholarships) - len(unique)
    print(f"  Removed {removed} duplicates")
    print(f"  Unique scholarships: {len(unique)}")

    # 5. Flag expired deadlines
    today = date.today()
    active = 0
    expired = 0
    unknown = 0
    for s in unique:
        dl = s.get("deadline", "Varies")
        try:
            dl_date = datetime.strptime(dl, "%Y-%m-%d").date()
            if dl_date < today:
                s["status"] = "EXPIRED"
                expired += 1
            else:
                s["status"] = "ACTIVE"
                active += 1
        except (ValueError, TypeError):
            s["status"] = "UNKNOWN"
            unknown += 1

    print(f"\n[5] Deadline analysis:")
    print(f"  Active (future deadline): {active}")
    print(f"  Expired (past deadline): {expired}")
    print(f"  Unknown (no parseable date): {unknown}")

    # 6. Write master CSV
    print(f"\n[6] Writing clean master CSV to: {OUTPUT_CSV}")
    fieldnames = ["id", "name", "criteria", "link", "deadline", "amount", "needBased", "status", "source_sheet"]
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for s in unique:
            writer.writerow(s)

    print(f"\n{'=' * 60}")
    print(f"DONE! {len(unique)} unique scholarships written to CSV.")
    print(f"  Active: {active} | Expired: {expired} | Unknown date: {unknown}")
    print(f"{'=' * 60}")

    return unique


if __name__ == "__main__":
    main()
